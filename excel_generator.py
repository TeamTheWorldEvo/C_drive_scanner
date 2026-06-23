# -*- coding: utf-8 -*-
"""Excel report generator for C Drive Scanner.

Generates a 3-sheet Excel report:
1. C盘清理总览 (Summary)
2. 可清理迁移详细清单 (Detailed List)
3. 操作指南 (Guide)
"""

from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from data_model import ScanResult, ScanItem, ActionType


class ExcelReportGenerator:
    """Generates the 3-sheet Excel report from ScanResult."""

    # Style constants
    TITLE_FONT = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    TITLE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    SECTION_FONT = Font(name="微软雅黑", size=12, bold=True, color="2F5496")
    SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    NORMAL_FONT = Font(name="微软雅黑", size=10)
    BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
    DELETE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    MIGRATE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    BOTH_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    SYSTEM_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    @classmethod
    def generate(cls, result: ScanResult, output_path: str):
        """Generate the complete 3-sheet Excel report."""
        wb = openpyxl.Workbook()
        cls._build_summary_sheet(wb, result)
        cls._build_details_sheet(wb, result)
        cls._build_guide_sheet(wb)
        wb.save(output_path)

    @classmethod
    def _apply_style(cls, cell, font=None, fill=None, alignment=None):
        cell.font = font or cls.NORMAL_FONT
        if fill:
            cell.fill = fill
        cell.alignment = alignment or cls.LEFT
        cell.border = cls.BORDER

    @classmethod
    def _fmt_size(cls, b: int) -> str:
        if b == 0:
            return "0"
        units = ["B", "KB", "MB", "GB", "TB"]
        size = float(b)
        unit_idx = 0
        while size >= 1024 and unit_idx < len(units) - 1:
            size /= 1024
            unit_idx += 1
        if unit_idx == 0:
            return f"{int(size)}{units[unit_idx]}"
        return f"{size:.1f}{units[unit_idx]}"

    # =========================================================================
    # Sheet 1: Summary
    # =========================================================================
    @classmethod
    def _build_summary_sheet(cls, wb, result: ScanResult):
        ws = wb.active
        ws.title = "C盘清理总览"

        # Title
        ws.merge_cells("A1:G1")
        cls._apply_style(ws["A1"], cls.TITLE_FONT, cls.TITLE_FILL, cls.CENTER)
        ws["A1"].value = "C盘清理分析报告"
        ws.row_dimensions[1].height = 45

        # Drive info
        ws.merge_cells("A2:G2")
        cls._apply_style(
            ws["A2"],
            Font(name="微软雅黑", size=10, color="FF0000", bold=True),
            PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            cls.CENTER
        )
        ws["A2"].value = (
            f"扫描时间: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}  |  "
            f"C盘容量: {result.total_drive_display} | "
            f"已使用: {result.used_drive_display} ({result.used_percent:.1f}%) | "
            f"可用: {result.free_drive_display}"
        )
        ws.row_dimensions[2].height = 30

        # Section header
        ws.merge_cells("A4:G4")
        cls._apply_style(ws["A4"], cls.SECTION_FONT, cls.SECTION_FILL, cls.LEFT)
        ws["A4"].value = "空间占用总览"
        ws.row_dimensions[4].height = 30

        # Compute category totals
        cat_data = {}
        for item in result.items:
            cat = item.category
            if cat not in cat_data:
                cat_data[cat] = {"size": 0, "count": 0, "action_types": set(),
                                  "main": "", "examples": []}
            cat_data[cat]["size"] += item.size_bytes
            cat_data[cat]["count"] += 1
            cat_data[cat]["action_types"].add(item.action_type)
            if len(cat_data[cat]["examples"]) < 3:
                cat_data[cat]["examples"].append(item.name)

        # Determine action type label
        def action_label(types):
            if ActionType.DELETE in types and ActionType.MIGRATE in types:
                return "删除+迁移"
            elif ActionType.DELETE in types:
                return "删除"
            elif ActionType.MIGRATE in types:
                return "迁移"
            elif ActionType.SYSTEM in types:
                return "系统优化"
            return "检查"

        # Determine main items text
        def main_items_text(examples, size):
            text = "、".join(examples[:3])
            if size > 10 * 1024**3:
                text = f"⚠ {text}"
            return text

        # Table header
        headers = ["类别", "空间占用", "占比(C盘)", "主要项目",
                    "建议操作", "项目数", "操作难度"]
        for j, h in enumerate(headers):
            cell = ws.cell(row=5, column=j+1, value=h)
            cls._apply_style(cell, cls.HEADER_FONT, cls.HEADER_FILL, cls.CENTER)
        ws.row_dimensions[5].height = 28

        # Sort categories by size
        sorted_cats = sorted(cat_data.items(), key=lambda x: x[1]["size"], reverse=True)

        row = 6
        for cat, data in sorted_cats:
            size_display = cls._fmt_size(data["size"])
            pct = (data["size"] / result.total_drive_bytes * 100) if result.total_drive_bytes else 0
            main_text = main_items_text(data["examples"], data["size"])
            act = action_label(data["action_types"])
            difficulty = "简单" if act in ("删除", "迁移") else "中等"

            values = [cat, size_display, f"{pct:.1f}%", main_text,
                      act, str(data["count"]), difficulty]
            for j, val in enumerate(values):
                cell = ws.cell(row=row, column=j+1, value=val)
                font = cls.NORMAL_FONT
                if j == 2 and pct > 5:
                    font = Font(name="微软雅黑", size=10, bold=True, color="FF0000")
                cls._apply_style(cell, font, alignment=cls.CENTER if j != 3 else cls.LEFT)
            ws.row_dimensions[row].height = 30
            row += 1

        # Total row
        total_size = sum(d["size"] for d in cat_data.values())
        total_pct = (total_size / result.total_drive_bytes * 100) if result.total_drive_bytes else 0
        total_count = sum(d["count"] for d in cat_data.values())
        total_values = ["合计", cls._fmt_size(total_size), f"{total_pct:.1f}%",
                        "以上所有可清理/迁移项目汇总", "—", str(total_count), "—"]
        for j, val in enumerate(total_values):
            cell = ws.cell(row=row, column=j+1, value=val)
            cls._apply_style(
                cell, cls.BOLD_FONT,
                PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
                cls.CENTER if j != 3 else cls.LEFT
            )
        ws.row_dimensions[row].height = 30
        row += 2

        # Tips
        cls._apply_style(
            ws.cell(row=row, column=1, value="重要提示"),
            Font(name="微软雅黑", size=12, bold=True, color="FF0000"),
            PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            cls.LEFT
        )
        ws.merge_cells(f"A{row}:G{row}")

        tips = [
            "1. 本工具只做扫描分析，不执行任何删除/迁移操作",
            "2. 标记为「删除」的文件/目录，建议先备份再执行操作",
            "3. 标记为「迁移」的文件/目录，可移动到D盘或其他磁盘",
            "4. hiberfil.sys 可通过管理员命令行执行 powercfg /h off 关闭休眠来释放",
            "5. WinSxS 可通过「磁盘清理→清理系统文件→Windows更新清理」安全清理",
            "6. pip/npm/NuGet 缓存可直接删除，下次构建会自动重新下载",
            "7. 操作前请确保已关闭相关应用程序",
        ]
        for tip in tips:
            row += 1
            ws.merge_cells(f"A{row}:G{row}")
            cls._apply_style(ws[f"A{row}"], cls.NORMAL_FONT, alignment=cls.LEFT)
            ws[f"A{row}"].value = tip
            ws.row_dimensions[row].height = 22

        # Disclaimer section
        row += 2
        cls._apply_style(
            ws.cell(row=row, column=1, value="免责声明与联系方式"),
            Font(name="微软雅黑", size=12, bold=True, color="2F5496"),
            PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
            cls.LEFT
        )
        ws.merge_cells(f"A{row}:G{row}")

        disclaimer_lines = [
            "免责声明：本工具仅供系统空间分析使用，仅提供扫描检测与Excel报告生成功能，不会执行任何实际的文件删除或迁移操作。",
            "用户根据本报告进行任何文件操作所产生的后果由用户自行承担。操作前请务必备份重要数据！",
            "联系方式：the_world_evo@163.com  |  GitHub: https://github.com/TeamTheWorldEvo/C_drive_scanner.git",
        ]
        for line in disclaimer_lines:
            row += 1
            ws.merge_cells(f"A{row}:G{row}")
            cls._apply_style(ws[f"A{row}"], cls.NORMAL_FONT, alignment=cls.LEFT)
            ws[f"A{row}"].value = line
            ws.row_dimensions[row].height = 22

        # Column widths
        widths = [18, 14, 14, 40, 14, 10, 14]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = w

    # =========================================================================
    # Sheet 2: Detailed List
    # =========================================================================
    @classmethod
    def _build_details_sheet(cls, wb, result: ScanResult):
        ws = wb.create_sheet("可清理迁移详细清单")

        # Title
        ws.merge_cells("A1:H1")
        cls._apply_style(ws["A1"], cls.TITLE_FONT, cls.TITLE_FILL, cls.CENTER)
        ws["A1"].value = "C盘可清理/迁移文件详细清单"
        ws.row_dimensions[1].height = 40

        ws.merge_cells("A2:H2")
        cls._apply_style(
            ws["A2"],
            Font(name="微软雅黑", size=9, color="333333"),
            PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            cls.CENTER
        )
        ws["A2"].value = (
            "绿色=可迁移 | 红色=可删除 | 橙色=建议迁移或删除 | 灰色=系统文件需谨慎"
        )

        # Headers
        headers = ["序号", "文件名/目录名", "存储位置", "大小", "类别",
                    "文件作用说明", "建议操作", "操作说明/注意事项"]
        for j, h in enumerate(headers):
            cell = ws.cell(row=4, column=j+1, value=h)
            cls._apply_style(cell, cls.HEADER_FONT, cls.HEADER_FILL, cls.CENTER)
        ws.row_dimensions[4].height = 30

        # Sort items by category then size
        sorted_items = sorted(result.items, key=lambda x: (x.category, -x.size_bytes))

        row = 5
        seq = 0
        last_category = None

        for item in sorted_items:
            # Category section header
            if item.category != last_category:
                last_category = item.category
                ws.merge_cells(f"A{row}:H{row}")
                cls._apply_style(
                    ws[f"A{row}"],
                    cls.SECTION_FONT, cls.SECTION_FILL, cls.LEFT
                )
                ws[f"A{row}"].value = f"▍{item.category}"
                ws.row_dimensions[row].height = 28
                row += 1

            seq += 1

            # Determine fill
            if item.action_type == ActionType.DELETE:
                fill = cls.DELETE_FILL
            elif item.action_type == ActionType.MIGRATE:
                fill = cls.MIGRATE_FILL
            elif item.action_type == ActionType.BOTH:
                fill = cls.BOTH_FILL
            elif item.action_type == ActionType.SYSTEM:
                fill = cls.SYSTEM_FILL
            else:
                fill = None

            values = [
                str(seq), item.name, item.path, item.size_display,
                item.category, item.description, item.suggested_action,
                item.notes,
            ]
            for j, val in enumerate(values):
                cell = ws.cell(row=row, column=j+1, value=val)
                font = cls.NORMAL_FONT
                if j == 0:
                    font = cls.BOLD_FONT
                elif j == 3:
                    font = Font(name="微软雅黑", size=10, bold=True)
                elif j == 6:
                    color = "C00000" if item.action_type == ActionType.DELETE else "006100"
                    font = Font(name="微软雅黑", size=10, bold=True, color=color)
                cls._apply_style(
                    cell, font, fill,
                    cls.CENTER if j in (0, 3, 4, 6) else cls.LEFT
                )
            ws.row_dimensions[row].height = 50
            row += 1

        # Column widths
        widths = [6, 28, 48, 8, 16, 45, 16, 42]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+1)].width = w

        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:H{row-1}"

    # =========================================================================
    # Sheet 3: Guide
    # =========================================================================
    @classmethod
    def _build_guide_sheet(cls, wb):
        ws = wb.create_sheet("操作指南")

        ws.merge_cells("A1:B1")
        cls._apply_style(ws["A1"], cls.TITLE_FONT, cls.TITLE_FILL, cls.CENTER)
        ws["A1"].value = "C盘清理操作指南"
        ws.row_dimensions[1].height = 40

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 80

        guide = [
            ("操作类型", "具体步骤"),
            ("🗑️ 清空回收站", "右键桌面「回收站」→「清空回收站」"),
            ("🗑️ 磁盘清理(基础)", "1. 打开「此电脑」\n2. 右键C盘→「属性」\n3. 点击「磁盘清理」\n4. 勾选所有可清理项→确定"),
            ("🗑️ 磁盘清理(系统文件)", "1. 在磁盘清理窗口中点击「清理系统文件」\n2. 勾选「Windows更新清理」\n3. 勾选「传递优化文件」\n4. 确定删除"),
            ("💤 关闭系统休眠(释放空间)", "1. 以管理员身份打开「命令提示符」\n2. 输入: powercfg /h off\n3. 按回车执行\n4. 如需恢复: powercfg /h on"),
            ("💾 调整虚拟内存", "1. 右键「此电脑」→「属性」→「高级系统设置」\n2. 「高级」→「性能」→「设置」→「高级」→「虚拟内存」\n3. 取消自动管理，C盘设为无分页文件\n4. D盘设为系统管理大小→确定→重启"),
            ("💬 迁移微信文件到D盘", "微信PC版(中文):\n1. 微信→「设置」→「文件管理」\n2. 更改路径到D:\\WeChat Files\n\n微信PC版(英文xwechat):\n1. 打开设置→更改存储路径\n2. 或手动迁移后创建符号链接"),
            ("🐍 清理pip缓存", "打开命令行执行:\npip cache purge\n\n或手动删除 AppData\\Local\\pip 目录"),
            ("📦 清理npm缓存", "打开命令行执行:\nnpm cache clean --force\n\n或手动删除 AppData\\Local\\npm-cache 目录"),
            ("🔧 清理NuGet缓存", "打开命令行执行:\ndotnet nuget locals all --clear"),
            ("🧹 清理Chrome缓存", "方法1: Chrome→设置→隐私和安全→清除浏览数据\n方法2: 删除 Chrome\\User Data\\Default\\Cache 目录"),
            ("🛠️ 清理PyCharm缓存", "方法1: PyCharm→File→Invalidate Caches\n方法2: 手动删除 JetBrains\\PyCharm2025.1\\caches"),
            ("📁 迁移用户文件夹", "对于Documents/Downloads等:\n1. 右键文件夹→「属性」→「位置」\n2. 点击「移动」→选择D盘目标位置\n3. 点击「应用」→确认移动文件"),
            ("🗑️ 删除临时文件", "方法1: 设置→系统→存储→临时文件\n方法2: Windows+R→输入 %temp% →全选删除\n方法3: 磁盘清理工具"),
            ("🗑️ 删除崩溃转储", "删除 AppData\\Local\\CrashDumps 目录下的.dmp文件"),
            ("📱 卸载不需要的应用", "设置→应用→应用和功能\n按大小排序，卸载不再使用的应用"),
            ("🐳 清理Docker", "1. docker system prune -a (清理未使用资源)\n2. Docker Desktop设置中更改镜像存储位置到D盘"),
            ("📝 检查虚拟机镜像", "检查VMware/VirtualBox/WSL虚拟磁盘文件(.vmdk/.vdi/.vhdx)\n如在C盘则迁移到D盘"),
        ]

        for i, (label, steps) in enumerate(guide):
            row = i + 3
            ws.row_dimensions[row].height = 22 if i == 0 else 80
            cell_a = ws.cell(row=row, column=1, value=label)
            cell_b = ws.cell(row=row, column=2, value=steps)
            if i == 0:
                cls._apply_style(cell_a, cls.HEADER_FONT, cls.HEADER_FILL, cls.CENTER)
                cls._apply_style(cell_b, cls.HEADER_FONT, cls.HEADER_FILL, cls.CENTER)
            else:
                cls._apply_style(cell_a, cls.NORMAL_FONT, alignment=cls.LEFT)
                cls._apply_style(cell_b, cls.NORMAL_FONT, alignment=cls.LEFT)
